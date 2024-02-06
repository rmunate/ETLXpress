#---------------------------------------------------------#
# Update Script Path                                      #
#---------------------------------------------------------#

from os.path import abspath, join, dirname
import sys

project_directory = abspath(join(dirname(__file__), '../../'))
if project_directory not in sys.path:
    sys.path.insert(0, project_directory)
    
    
#---------------------------------------------------------#
# Class For Excel File                                    #
# Last Modification: 05-02-2024                           #
# Developers:                                             #
# -------------    Raul Mauricio Uñate Castro             #
#---------------------------------------------------------#

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from core.Clarity.Logger import Logger

class Excel:
    
    # Initialize the workbook and active sheet
    def __init__(self):
        
        # Log Registration
        Logger().info(f"Excel Class: Initialized")
        
        # Start Workbook
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active

    # Set the background color for headers
    def setBackgroundColorHeader(self, color):
        
        # Log Registration
        Logger().info(f"Excel Class: Set Background Color Header <{color}>")
        
        # Save Values
        self.headers_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
    # Set bold font style for headers
    def setTextBoldHeader(self, bold = False):
        
        # Log Registration
        Logger().info(f"Excel Class: Set Bold Header <{bold}>")
        
        # Save Values
        self.headers_style = Font(bold=bold)
        
     # Set text color for headers
    def setTextColorHeader(self, color):
        
        # Log Registration
        Logger().info(f"Excel Class: Set Text Color Header <{color}>")
        
        # Save Values
        self.headers_style.color = color
        
    # Set the sheet name
    def setSheetName(self, sheet_name = 'ETXLpress'):
        
        # Log Registration
        Logger().info(f"Excel Class: Set Sheet Name <{sheet_name}>")
        
        # Save Values
        self.sheet.title = sheet_name

    # Write headers to the Excel sheet with specified styles
    def headers(self, headers):
        
        # Log Registration
        Logger().info(f"Excel Class: Set Headers")
        
        # Set Values
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = self.sheet[f'{col_letter}1']
            cell.value = header
            cell.font = self.headers_style
            cell.fill = self.headers_fill

    # Write data rows to the Excel sheet
    def rows(self, data):
        
        # Log Registration
        Logger().info(f"Excel Class: Set Rows")
        
        # Set Values
        for row in data:
            self.sheet.append(row)

    # Save the Excel workbook to a file
    def save(self, filename = 'ETLXpress_Report.xlsx'):
        
        # Log Registration
        Logger().info(f"Excel Class: Save File <{filename}>")
        
        # Save Excel
        self.workbook.save(filename)
